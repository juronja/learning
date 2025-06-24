import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
products = inv_file["products"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}

for row in range(2, products.max_row + 1):
    supplier = products.cell(row, 4).value
    inventory = products.cell(row, 2).value
    price = products.cell(row, 3).value
    product_number = products.cell(row, 1).value
    total_price = products.cell(row, 5)

    # calculate number of products per supplier
    if supplier in products_per_supplier:
        products_per_supplier[supplier] = products_per_supplier[supplier] + 1
    else:
        products_per_supplier[supplier] = 1

    # calculate total value of inventory per supplier
    if supplier in total_value_per_supplier:
        total_value_per_supplier[supplier] = total_value_per_supplier[supplier] + inventory * price
    else:
        total_value_per_supplier[supplier] = inventory * price

    # products with inventory less than 10
    if inventory < 10:
        products_under_10[product_number] = inventory

    # add total price column
    total_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10)

inv_file.save("inventory_with_total_value.xlsx")


